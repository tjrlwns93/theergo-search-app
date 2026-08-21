# -*- coding: utf-8 -*-
"""
더에르고 검색 웹앱 (Streamlit)
- 비밀번호 로그인 후에만 화면이 보인다.
- [검색어]를 넣으면 theergo.co.kr 첫 번째 결과의 제목/링크를 가져온다.
- 결과를 Supabase(PostgreSQL)에 저장하고, 저장 이력을 보여준다.
- 접속정보/비밀번호는 코드가 아니라 secrets(.streamlit/secrets.toml)에서 읽는다.
"""
import base64
import datetime
import html as html_mod
import io
import json
import re
import time
import urllib.parse
import urllib.request

import openpyxl
import psycopg2
import streamlit as st
import streamlit.components.v1 as components

# ──────────────────────────────────────────────────────────────
# 설정 (secrets 에서 읽음 — 코드에 비밀정보를 두지 않는다)
# ──────────────────────────────────────────────────────────────
BASE = "https://theergo.co.kr"
SEARCH_URL = BASE + "/product/search.html?keyword="
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Search-Bot"}

# 연습용 주문 API (인증 불필요, date/page 파라미터, 100건씩 페이지네이션)
ORDERS_API = "https://sp.ermore.co.kr/api/edu-leader/practice/orders"

# 날씨 API (Open-Meteo, 무료·키 불필요) — 서울 기준
WEATHER_API = "https://api.open-meteo.com/v1/forecast"
SEOUL_LAT, SEOUL_LON = 37.5665, 126.9780

st.set_page_config(page_title="더에르고 검색기", page_icon="🔎")


def get_secret(key: str, default=None):
    try:
        return st.secrets[key]
    except Exception:
        return default


# ──────────────────────────────────────────────────────────────
# 1) 비밀번호 로그인 게이트
# ──────────────────────────────────────────────────────────────
def check_password() -> bool:
    if st.session_state.get("auth_ok"):
        return True

    st.title("🔒 로그인")
    st.caption("비밀번호를 입력해야 화면이 보입니다.")
    pw = st.text_input("비밀번호", type="password")
    if st.button("로그인"):
        real = get_secret("app_password")
        if real and pw == real:
            st.session_state["auth_ok"] = True
            st.rerun()
        else:
            st.error("비밀번호가 틀렸습니다.")
    return False


# ──────────────────────────────────────────────────────────────
# 2) 데이터베이스 (Supabase PostgreSQL)
# ──────────────────────────────────────────────────────────────
def get_conn():
    # 방법 1 (권장): 항목별 접속정보 — 비번 특수문자/URL 인코딩 걱정 없음
    host = get_secret("db_host")
    if host:
        return psycopg2.connect(
            host=host,
            port=int(get_secret("db_port", 5432)),
            user=get_secret("db_user"),
            password=get_secret("db_password"),
            dbname=get_secret("db_name", "postgres"),
            sslmode="require",
            connect_timeout=15,
        )

    # 방법 2: 전체 연결 URL (특수문자는 직접 URL 인코딩 필요)
    url = get_secret("db_url")
    if not url:
        raise RuntimeError(
            "secrets 에 db_host(항목별) 또는 db_url 이 필요합니다."
        )
    if "<region>" in url or "[YOUR-PASSWORD]" in url or "[" in url:
        raise RuntimeError(
            "db_url 에 <region> / [YOUR-PASSWORD] 같은 자리표시자가 그대로 있습니다. "
            "실제 값으로 바꿔주세요."
        )
    if "sslmode=" not in url:
        url += ("&" if "?" in url else "?") + "sslmode=require"
    return psycopg2.connect(url, connect_timeout=15)


def init_db():
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS search_results (
                id BIGSERIAL PRIMARY KEY,
                keyword TEXT NOT NULL,
                title TEXT,
                link TEXT,
                searched_at TIMESTAMPTZ NOT NULL DEFAULT now()
            );
            """
        )
        # 주문 테이블: order_no 를 PK 로 두어 같은 날짜 재수집 시 중복 방지
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS orders (
                order_no       TEXT PRIMARY KEY,
                order_date     DATE NOT NULL,
                ordered_at     TEXT,
                product        TEXT,
                product_option TEXT,
                qty            INTEGER,
                amount         BIGINT,
                status         TEXT,
                collected_at   TIMESTAMPTZ NOT NULL DEFAULT now()
            );
            """
        )
        cur.execute(
            "CREATE INDEX IF NOT EXISTS idx_orders_date ON orders (order_date);"
        )
        # 날씨 테이블: (도시, 날짜) 를 PK 로 두어 같은 날짜 재수집 시 중복 방지
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS weather (
                city         TEXT NOT NULL DEFAULT 'Seoul',
                weather_date DATE NOT NULL,
                temp_max     REAL,
                temp_min     REAL,
                precipitation REAL,
                collected_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                PRIMARY KEY (city, weather_date)
            );
            """
        )
        # 카페24 OAuth 토큰 저장 (몰당 1행). 고객정보는 저장하지 않음.
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS cafe24_token (
                mall_id                  TEXT PRIMARY KEY,
                access_token             TEXT,
                refresh_token            TEXT,
                expires_at               TEXT,
                refresh_token_expires_at TEXT,
                updated_at               TIMESTAMPTZ NOT NULL DEFAULT now()
            );
            """
        )
        conn.commit()


def save_results(rows):
    """rows: list of (keyword, title, link)"""
    with get_conn() as conn, conn.cursor() as cur:
        cur.executemany(
            "INSERT INTO search_results (keyword, title, link) VALUES (%s, %s, %s)",
            rows,
        )
        conn.commit()


def fetch_history(limit=200):
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT keyword, title, link, searched_at "
            "FROM search_results ORDER BY id DESC LIMIT %s",
            (limit,),
        )
        return cur.fetchall()


# ──────────────────────────────────────────────────────────────
# 2-1) 주문 수집 (연습용 API → DB)
# ──────────────────────────────────────────────────────────────
def fetch_orders(date_str, progress=None, max_pages=1000):
    """해당 날짜의 모든 주문을 page 를 넘기며(100건씩) 모아서 반환.
    반환: (rows_list, total). 종료조건: 빈 페이지 또는 total 도달."""
    all_rows = []
    total = None
    page = 1
    while page <= max_pages:
        url = ORDERS_API + "?" + urllib.parse.urlencode({"date": date_str, "page": page})
        req = urllib.request.Request(
            url, headers={"User-Agent": HEADERS["User-Agent"], "Accept": "application/json"}
        )
        with urllib.request.urlopen(req, timeout=25) as resp:
            data = json.loads(resp.read().decode("utf-8", "replace"))
        rows = data.get("rows") or []
        if not rows:
            break
        all_rows.extend(rows)
        total = data.get("total")
        if progress:
            progress(len(all_rows), total, page)
        if total is not None and len(all_rows) >= total:
            break
        page += 1
    return all_rows, total


def upsert_orders(date_obj, rows):
    """order_no 기준 UPSERT. 같은 주문을 다시 넣어도 중복으로 쌓이지 않고 갱신됨.
    반환: (신규 건수, 갱신 건수)."""
    if not rows:
        return 0, 0
    new_cnt = upd_cnt = 0
    with get_conn() as conn, conn.cursor() as cur:
        for r in rows:
            cur.execute(
                """
                INSERT INTO orders
                    (order_no, order_date, ordered_at, product, product_option,
                     qty, amount, status, collected_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, now())
                ON CONFLICT (order_no) DO UPDATE SET
                    order_date     = EXCLUDED.order_date,
                    ordered_at     = EXCLUDED.ordered_at,
                    product        = EXCLUDED.product,
                    product_option = EXCLUDED.product_option,
                    qty            = EXCLUDED.qty,
                    amount         = EXCLUDED.amount,
                    status         = EXCLUDED.status,
                    collected_at   = now()
                RETURNING (xmax = 0) AS inserted
                """,
                (
                    r.get("order_no"), date_obj, r.get("ordered_at"),
                    r.get("product"), r.get("option"), r.get("qty"),
                    r.get("amount"), r.get("status"),
                ),
            )
            if cur.fetchone()[0]:
                new_cnt += 1
            else:
                upd_cnt += 1
        conn.commit()
    return new_cnt, upd_cnt


def fetch_orders_by_date(date_obj, limit=2000):
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT order_no, ordered_at, product, product_option, qty, amount, status "
            "FROM orders WHERE order_date = %s ORDER BY order_no LIMIT %s",
            (date_obj, limit),
        )
        return cur.fetchall()


# ──────────────────────────────────────────────────────────────
# 2-2) 날씨 수집 (Open-Meteo → DB), 서울 기준
# ──────────────────────────────────────────────────────────────
def fetch_weather(start_date, end_date):
    """서울의 날짜별 최고/최저기온·강수량을 가져와 리스트로 반환.
    반환: [(date, tmax, tmin, precip), ...]"""
    url = WEATHER_API + "?" + urllib.parse.urlencode({
        "latitude": SEOUL_LAT,
        "longitude": SEOUL_LON,
        "daily": "temperature_2m_max,temperature_2m_min,precipitation_sum",
        "timezone": "Asia/Seoul",
        "start_date": start_date,
        "end_date": end_date,
    })
    req = urllib.request.Request(
        url, headers={"User-Agent": HEADERS["User-Agent"], "Accept": "application/json"}
    )
    with urllib.request.urlopen(req, timeout=25) as resp:
        data = json.loads(resp.read().decode("utf-8", "replace"))
    daily = data.get("daily") or {}
    times = daily.get("time") or []
    tmax = daily.get("temperature_2m_max") or []
    tmin = daily.get("temperature_2m_min") or []
    prcp = daily.get("precipitation_sum") or []
    rows = []
    for i, d in enumerate(times):
        rows.append((
            d,
            tmax[i] if i < len(tmax) else None,
            tmin[i] if i < len(tmin) else None,
            prcp[i] if i < len(prcp) else None,
        ))
    return rows


def upsert_weather(rows, city="Seoul"):
    """(도시,날짜) 기준 UPSERT. 같은 날짜를 다시 가져와도 중복 없이 갱신.
    반환: (신규 건수, 갱신 건수)."""
    if not rows:
        return 0, 0
    new_cnt = upd_cnt = 0
    with get_conn() as conn, conn.cursor() as cur:
        for d, tmax, tmin, prcp in rows:
            cur.execute(
                """
                INSERT INTO weather
                    (city, weather_date, temp_max, temp_min, precipitation, collected_at)
                VALUES (%s, %s, %s, %s, %s, now())
                ON CONFLICT (city, weather_date) DO UPDATE SET
                    temp_max      = EXCLUDED.temp_max,
                    temp_min      = EXCLUDED.temp_min,
                    precipitation = EXCLUDED.precipitation,
                    collected_at  = now()
                RETURNING (xmax = 0) AS inserted
                """,
                (city, d, tmax, tmin, prcp),
            )
            if cur.fetchone()[0]:
                new_cnt += 1
            else:
                upd_cnt += 1
        conn.commit()
    return new_cnt, upd_cnt


def fetch_weather_saved(city="Seoul", limit=365):
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT weather_date, temp_max, temp_min, precipitation "
            "FROM weather WHERE city = %s ORDER BY weather_date DESC LIMIT %s",
            (city, limit),
        )
        return cur.fetchall()


def fetch_weather_collected_today(city="Seoul"):
    """오늘(한국시간) 수집한 날씨만 반환."""
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT weather_date, temp_max, temp_min, precipitation "
            "FROM weather WHERE city = %s "
            "AND (collected_at AT TIME ZONE 'Asia/Seoul')::date "
            "    = (now() AT TIME ZONE 'Asia/Seoul')::date "
            "ORDER BY weather_date",
            (city,),
        )
        return cur.fetchall()


# ──────────────────────────────────────────────────────────────
# 2-3) 노션 연동 (Open-Meteo 로 모은 날씨를 노션 DB 에 한 줄씩 추가)
# ──────────────────────────────────────────────────────────────
NOTION_API = "https://api.notion.com/v1"
NOTION_VERSION = "2022-06-28"


def add_weather_rows_to_notion(rows):
    """rows: [(date, tmax, tmin, precip), ...] 를 노션 DB 에 한 줄씩 추가.
    '테스트'(title)=날씨 내용, '담당자'(people)=secrets 의 notion_person_id.
    반환: 추가한 줄 수."""
    token = get_secret("notion_token")
    db_id = get_secret("notion_db_id")
    person_id = get_secret("notion_person_id")
    if not token or not db_id:
        raise RuntimeError("secrets 에 notion_token / notion_db_id 가 필요합니다.")
    headers = {
        "Authorization": f"Bearer {token}",
        "Notion-Version": NOTION_VERSION,
        "Content-Type": "application/json",
    }
    added = 0
    for d, tmax, tmin, prcp in rows:
        content = f"{d} 서울 · 최고 {tmax}°C / 최저 {tmin}°C / 강수 {prcp}mm"
        props = {"테스트": {"title": [{"text": {"content": content}}]}}
        if person_id:
            props["담당자"] = {"people": [{"id": person_id}]}
        payload = {"parent": {"database_id": db_id}, "properties": props}
        req = urllib.request.Request(
            NOTION_API + "/pages",
            data=json.dumps(payload).encode("utf-8"),
            headers=headers,
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=25) as resp:
            resp.read()
        added += 1
    return added


# ──────────────────────────────────────────────────────────────
# 2-4) 카페24 관리자 API (OAuth + 토큰 자동갱신). 고객정보 미수집.
# ──────────────────────────────────────────────────────────────
CAFE24_API_VERSION = "2022-06-01"  # X-Cafe24-Api-Version (앱 설정과 맞추세요)


def _cafe24_cfg():
    return (
        get_secret("cafe24_mall_id"),
        get_secret("cafe24_client_id"),
        get_secret("cafe24_client_secret"),
        get_secret("cafe24_redirect_uri"),
    )


def cafe24_authorize_url(state="ergo2-state"):
    """1단계: 승인 요청 주소. scope 는 주문 읽기만(고객정보 scope 없음)."""
    mall_id, client_id, _, redirect_uri = _cafe24_cfg()
    params = {
        "response_type": "code",
        "client_id": client_id,
        "state": state,
        "redirect_uri": redirect_uri,
        "scope": "mall.read_order",
    }
    return (
        f"https://{mall_id}.cafe24api.com/api/v2/oauth/authorize?"
        + urllib.parse.urlencode(params)
    )


def _cafe24_save_token(mall_id, tok):
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO cafe24_token
                (mall_id, access_token, refresh_token, expires_at,
                 refresh_token_expires_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, now())
            ON CONFLICT (mall_id) DO UPDATE SET
                access_token             = EXCLUDED.access_token,
                refresh_token            = EXCLUDED.refresh_token,
                expires_at               = EXCLUDED.expires_at,
                refresh_token_expires_at = EXCLUDED.refresh_token_expires_at,
                updated_at               = now()
            """,
            (
                mall_id, tok.get("access_token"), tok.get("refresh_token"),
                tok.get("expires_at"), tok.get("refresh_token_expires_at"),
            ),
        )
        conn.commit()


def _cafe24_load_token():
    mall_id = get_secret("cafe24_mall_id")
    with get_conn() as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT access_token, refresh_token, expires_at "
            "FROM cafe24_token WHERE mall_id = %s",
            (mall_id,),
        )
        row = cur.fetchone()
    if not row:
        return None
    return {"access_token": row[0], "refresh_token": row[1], "expires_at": row[2]}


def _cafe24_token_request(form):
    """POST /oauth/token (Basic 인증). 응답 토큰을 DB에 저장하고 반환."""
    mall_id, client_id, client_secret, _ = _cafe24_cfg()
    url = f"https://{mall_id}.cafe24api.com/api/v2/oauth/token"
    basic = base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()
    req = urllib.request.Request(
        url,
        data=urllib.parse.urlencode(form).encode(),
        method="POST",
        headers={
            "Authorization": f"Basic {basic}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
    )
    with urllib.request.urlopen(req, timeout=25) as resp:
        tok = json.loads(resp.read().decode("utf-8"))
    _cafe24_save_token(mall_id, tok)
    return tok


def cafe24_exchange_code(code):
    """3~4단계: 코드 → 액세스/갱신 토큰 교환."""
    _, _, _, redirect_uri = _cafe24_cfg()
    return _cafe24_token_request({
        "grant_type": "authorization_code",
        "code": code,
        "redirect_uri": redirect_uri,
    })


def cafe24_refresh_token():
    """갱신 토큰으로 새 액세스 토큰 발급(자동 갱신용)."""
    tok = _cafe24_load_token()
    if not tok or not tok.get("refresh_token"):
        raise RuntimeError("갱신할 refresh_token 이 없습니다. 다시 연결(승인)하세요.")
    return _cafe24_token_request({
        "grant_type": "refresh_token",
        "refresh_token": tok["refresh_token"],
    })


def cafe24_get(path, params=None):
    """액세스 토큰으로 GET. 401(만료)이면 자동 갱신 후 1회 재시도."""
    mall_id = get_secret("cafe24_mall_id")
    tok = _cafe24_load_token()
    if not tok:
        raise RuntimeError("카페24가 아직 연결되지 않았습니다. 먼저 연결(승인)하세요.")

    def _call(access_token):
        url = f"https://{mall_id}.cafe24api.com{path}"
        if params:
            url += "?" + urllib.parse.urlencode(params)
        req = urllib.request.Request(url, headers={
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
            "X-Cafe24-Api-Version": CAFE24_API_VERSION,
        })
        with urllib.request.urlopen(req, timeout=25) as resp:
            return json.loads(resp.read().decode("utf-8"))

    try:
        return _call(tok["access_token"])
    except urllib.error.HTTPError as e:
        if e.code == 401:  # 액세스 토큰 만료 → 자동 갱신 후 재시도
            newtok = cafe24_refresh_token()
            return _call(newtok["access_token"])
        raise


def cafe24_today_summary():
    """오늘 주문 건수와 매출(결제금액 합계). 금액·건수만 요청 — 고객정보 미수집."""
    today = datetime.date.today().strftime("%Y-%m-%d")
    cnt = cafe24_get(
        "/api/v2/admin/orders/count",
        {"start_date": today, "end_date": today},
    ).get("count", 0)

    total = 0.0
    offset, limit = 0, 100
    while True:
        data = cafe24_get("/api/v2/admin/orders", {
            "start_date": today, "end_date": today,
            "limit": limit, "offset": offset,
            # 금액 필드만 요청 — 이름/연락처/주소는 애초에 받지 않음
            "fields": "order_id,actual_order_amount,payment_amount",
        })
        orders = data.get("orders") or []
        if not orders:
            break
        for o in orders:
            amt = o.get("actual_order_amount") or o.get("payment_amount") or 0
            try:
                total += float(amt)
            except (TypeError, ValueError):
                pass
        if len(orders) < limit:
            break
        offset += limit
    return cnt, total


# ──────────────────────────────────────────────────────────────
# 3) 검색 (theergo.co.kr)
# ──────────────────────────────────────────────────────────────
def fetch(keyword: str) -> str:
    url = SEARCH_URL + urllib.parse.quote(keyword)
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=25) as resp:
        return resp.read().decode("utf-8", "replace")


def strip_tags(fragment: str) -> str:
    text = re.sub(r"<[^>]+>", " ", fragment)
    text = html_mod.unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"^상품명\s*:\s*", "", text)
    return text


def parse_first_result(page_html: str):
    start = page_html.find("prdList")
    if start == -1:
        return None, None
    region = page_html[start:]
    m = re.search(
        r'<strong class="name">\s*<a\s+href="([^"]+)"[^>]*>(.*?)</a>',
        region,
        re.S,
    )
    if not m:
        return None, None
    href = html_mod.unescape(m.group(1).strip())
    title = strip_tags(m.group(2))
    link = urllib.parse.urljoin(BASE, href)
    return (title or None), link


def search_one(keyword: str):
    try:
        title, link = parse_first_result(fetch(keyword))
        if title is None:
            return "검색결과 없음", ""
        return title, link
    except Exception as e:
        return f"오류: {e}", ""


def to_excel_bytes(rows, header):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "검색결과"
    ws.append(header)
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────
# 4) 메인 화면
# ──────────────────────────────────────────────────────────────
def render_search():
    st.write("검색어를 넣으면 theergo.co.kr 첫 번째 결과의 제목·링크를 가져와 저장합니다.")

    st.subheader("1. 검색어 입력")
    tab1, tab2 = st.tabs(["직접 입력", "엑셀 업로드"])

    keywords = []
    with tab1:
        text = st.text_area(
            "한 줄에 하나씩 입력",
            value="잠옷\n매트\n치약\n스티커",
            height=150,
        )
        if text.strip():
            keywords = [ln.strip() for ln in text.splitlines() if ln.strip()]

    with tab2:
        up = st.file_uploader("[검색어] 열이 있는 .xlsx", type=["xlsx"])
        col = st.text_input("검색어 열 이름", value="검색어")
        if up is not None:
            wb = openpyxl.load_workbook(up)
            ws = wb.active
            header = [c.value for c in ws[1]]
            if col in header:
                idx = header.index(col)
                keywords = [
                    str(row[idx]).strip()
                    for row in ws.iter_rows(min_row=2, values_only=True)
                    if idx < len(row) and row[idx] and str(row[idx]).strip()
                ]
                st.success(f"{len(keywords)}개 검색어를 읽었습니다: {keywords}")
            else:
                st.error(f"'{col}' 열이 없습니다. 헤더: {header}")

    st.subheader("2. 검색 실행")
    if st.button("🚀 검색하고 저장하기", type="primary", disabled=not keywords):
        rows = []
        prog = st.progress(0.0)
        area = st.empty()
        for i, kw in enumerate(keywords, 1):
            title, link = search_one(kw)
            rows.append((kw, title, link))
            area.write(f"[{i}/{len(keywords)}] **{kw}** → {title}")
            prog.progress(i / len(keywords))
            if i < len(keywords):
                time.sleep(1.0)
        try:
            save_results(rows)
            st.success(f"완료! {len(rows)}건을 Supabase에 저장했습니다.")
        except Exception as e:
            st.error(f"저장 실패: {e}")
        st.dataframe(
            [{"검색어": r[0], "제목": r[1], "링크": r[2]} for r in rows],
            use_container_width=True,
        )
        st.download_button(
            "⬇️ 엑셀로 내려받기",
            data=to_excel_bytes(rows, ["검색어", "제목", "링크"]),
            file_name="검색결과.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.subheader("3. 저장 이력 (최근 200건)")
    if st.button("🔄 이력 새로고침"):
        st.session_state["_reload"] = True
    try:
        hist = fetch_history()
        if hist:
            st.dataframe(
                [
                    {
                        "검색어": h[0],
                        "제목": h[1],
                        "링크": h[2],
                        "시각": h[3].strftime("%Y-%m-%d %H:%M:%S"),
                    }
                    for h in hist
                ],
                use_container_width=True,
            )
        else:
            st.info("아직 저장된 이력이 없습니다.")
    except Exception as e:
        st.error(f"이력 조회 실패: {e}")


def render_orders():
    st.write("날짜를 고르고 버튼을 누르면 그 날 주문을 API에서 100건씩 모두 가져와 DB에 쌓습니다.")

    picked = st.date_input("수집할 날짜", value=datetime.date(2026, 8, 18))
    date_str = picked.strftime("%Y-%m-%d")

    if st.button("📦 이 날짜 주문 수집", type="primary"):
        prog = st.progress(0.0)
        area = st.empty()

        def on_prog(got, total, page):
            if total:
                prog.progress(min(got / total, 1.0))
            area.write(f"{page}페이지까지 {got}/{total or '?'}건 가져오는 중…")

        try:
            rows, total = fetch_orders(date_str, progress=on_prog)
        except Exception as e:
            st.error(f"API 호출 실패: {e}")
            st.stop()

        prog.progress(1.0)
        if not rows:
            st.info(f"{date_str} 에는 주문이 없습니다.")
        else:
            try:
                new_cnt, upd_cnt = upsert_orders(picked, rows)
                st.success(
                    f"완료! {date_str} 주문 {len(rows)}건 수집 "
                    f"→ 신규 {new_cnt}건 · 갱신 {upd_cnt}건 (중복은 자동 제외)"
                )
            except Exception as e:
                st.error(f"저장 실패: {e}")

    st.divider()
    st.caption(f"DB에 저장된 {date_str} 주문")
    try:
        saved = fetch_orders_by_date(picked)
        st.write(f"저장된 주문: **{len(saved)}건**")
        if saved:
            st.dataframe(
                [
                    {
                        "주문번호": s[0],
                        "주문시각": s[1],
                        "상품": s[2],
                        "옵션": s[3],
                        "수량": s[4],
                        "금액": s[5],
                        "상태": s[6],
                    }
                    for s in saved
                ],
                use_container_width=True,
            )
    except Exception as e:
        st.error(f"주문 조회 실패: {e}")


def render_weather():
    st.write("서울의 날짜별 최고·최저 기온과 강수량을 Open-Meteo에서 가져와 DB에 저장합니다.")

    today = datetime.date.today()
    c1, c2 = st.columns(2)
    with c1:
        start = st.date_input("시작 날짜", value=today - datetime.timedelta(days=6))
    with c2:
        end = st.date_input("끝 날짜", value=today)

    if start > end:
        st.warning("시작 날짜가 끝 날짜보다 늦습니다.")
    elif st.button("🌤️ 날씨 가져와 저장", type="primary"):
        try:
            rows = fetch_weather(start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"))
        except Exception as e:
            st.error(f"API 호출 실패: {e}")
            st.stop()
        if not rows:
            st.info("해당 기간의 날씨 데이터가 없습니다.")
        else:
            try:
                new_cnt, upd_cnt = upsert_weather(rows)
                st.success(
                    f"완료! {len(rows)}일치 수집 "
                    f"→ 신규 {new_cnt}건 · 갱신 {upd_cnt}건 (중복은 자동 제외)"
                )
            except Exception as e:
                st.error(f"저장 실패: {e}")

    st.divider()
    st.caption("DB에 저장된 서울 날씨 (최근순)")
    try:
        saved = fetch_weather_saved()
        st.write(f"저장된 날짜: **{len(saved)}일**")
        if saved:
            table = [
                {
                    "날짜": s[0].strftime("%Y-%m-%d"),
                    "최고기온(°C)": s[1],
                    "최저기온(°C)": s[2],
                    "강수량(mm)": s[3],
                }
                for s in saved
            ]
            st.dataframe(table, use_container_width=True)
    except Exception as e:
        st.error(f"날씨 조회 실패: {e}")

    # 노션으로 보내기 — 오늘 모은 날씨를 노션 DB에 한 줄씩 추가
    if get_secret("notion_token"):
        st.divider()
        st.caption("노션으로 보내기 ('테스트' 칸=날씨 내용, '담당자' 칸=지정한 사용자)")
        if st.button("📤 오늘 모은 서울 날씨를 노션에 한 줄씩 추가"):
            try:
                today_rows = fetch_weather_collected_today()
            except Exception as e:
                st.error(f"조회 실패: {e}")
                today_rows = []
            if not today_rows:
                st.info("오늘 수집한 날씨가 없습니다. 위에서 먼저 수집하세요.")
            else:
                rows = [
                    (r[0].strftime("%Y-%m-%d"), r[1], r[2], r[3]) for r in today_rows
                ]
                try:
                    n = add_weather_rows_to_notion(rows)
                    st.success(f"노션에 {n}줄 추가했습니다.")
                except Exception as e:
                    st.error(f"노션 추가 실패: {e}")


def render_cafe24():
    st.write(
        "카페24(몰: **ergo2**) 관리자 API로 **오늘 매출·주문 건수**만 봅니다. "
        "고객 이름·연락처·주소는 가져오지 않습니다(scope: 주문 읽기, 금액·건수만 집계)."
    )
    mall_id, client_id, client_secret, redirect_uri = _cafe24_cfg()
    if not (mall_id and client_id and client_secret and redirect_uri):
        st.warning(
            "secrets 에 cafe24_client_id / cafe24_client_secret / "
            "cafe24_redirect_uri 를 넣어주세요. (몰 아이디는 ergo2로 설정됨)"
        )
        return

    # OAuth 콜백: 승인 후 ?code= 로 돌아오면 토큰 교환
    code = st.query_params.get("code")
    if code:
        try:
            cafe24_exchange_code(code)
            st.query_params.clear()
            st.success("카페24 연결 완료! 토큰을 저장했습니다.")
        except Exception as e:
            st.error(f"토큰 교환 실패: {e}")

    try:
        tok = _cafe24_load_token()
    except Exception as e:
        st.error(f"토큰 조회 실패: {e}")
        return

    if not tok:
        st.info("아직 연결 전입니다. 아래 버튼으로 승인하세요.")
        st.link_button("🔗 카페24 연결(승인)", cafe24_authorize_url())
        return

    st.success("연결됨 · 요청 시 토큰이 만료됐으면 자동으로 갱신됩니다.")
    if st.button("📊 오늘 매출·주문 건수 보기", type="primary"):
        try:
            cnt, total = cafe24_today_summary()
            c1, c2 = st.columns(2)
            c1.metric("오늘 주문 건수", f"{cnt:,} 건")
            c2.metric("오늘 매출(결제금액 합계)", f"{int(total):,} 원")
        except Exception as e:
            st.error(f"조회 실패: {e}")
    if st.button("🔄 토큰 수동 갱신"):
        try:
            cafe24_refresh_token()
            st.success("토큰을 갱신했습니다.")
        except Exception as e:
            st.error(f"갱신 실패: {e}")


# ──────────────────────────────────────────────────────────────
# 5) 모바일 공통 규칙 (한 곳에서 전체 화면에 적용)
# ──────────────────────────────────────────────────────────────
def mobile_css():
    st.markdown(
        """
        <style>
        /* 페이지가 좌우로 밀리지 않게 */
        html, body, [data-testid="stAppViewContainer"], .stApp { overflow-x: hidden !important; max-width: 100%; }
        /* 표: 넓으면 표 "안에서만" 가로 스크롤 (페이지는 안 밀림) */
        [data-testid="stDataFrame"], [data-testid="stDataFrameResizable"] { max-width: 100% !important; }
        [data-testid="stTable"], .stMarkdown table { display:block; overflow-x:auto; max-width:100%; }
        /* 입력칸 글자 16px 이상 (iOS 자동 확대 방지) */
        input, textarea, select,
        [data-baseweb="input"] input, [data-baseweb="textarea"] textarea, [data-baseweb="base-input"] input { font-size: 16px !important; }
        /* 버튼: 손가락으로 눌릴 만큼 크게 */
        .stButton>button, .stDownloadButton>button, [data-testid="stFormSubmitButton"] button,
        [data-testid="stLinkButton"] a { min-height: 44px !important; font-size: 16px !important; }
        /* 긴 숫자·금액이 줄바꿈되어 깨지지 않게 */
        [data-testid="stMetricValue"] { white-space: nowrap !important; }
        /* 팝오버/다이얼로그/서랍(사이드바): 고정너비 금지, 화면에 맞게 */
        [data-baseweb="popover"] > div, div[role="dialog"], [data-testid="stDialog"] > div { max-width: 96vw !important; }
        section[data-testid="stSidebar"] { max-width: 85vw !important; }
        @media (max-width: 480px){
          .block-container { padding-left: .8rem !important; padding-right: .8rem !important; padding-top: 2.6rem !important; }
          [data-testid="stMetricValue"] { font-size: 1.5rem !important; }
          section[data-testid="stSidebar"] { width: 82vw !important; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def pwa_component():
    """PWA: manifest/아이콘 head 주입 + 옛 캐시 제거 + '설치' 버튼(안드로이드 즉시설치 / iOS 안내)."""
    components.html(
        """
<style>
  * { box-sizing: border-box; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; }
  #bar { text-align: right; }
  #installBtn {
    display:none; border:0; background:#2F6FED; color:#fff; font-size:15px; font-weight:600;
    padding:10px 16px; border-radius:10px; min-height:44px; cursor:pointer;
    box-shadow:0 2px 8px rgba(47,111,237,.35);
  }
  #ios { display:none; position:fixed; inset:0; background:rgba(0,0,0,.55);
         align-items:center; justify-content:center; padding:20px; z-index:99999; }
  #ios .card { background:#fff; border-radius:16px; max-width:320px; width:100%; padding:20px; text-align:center; }
  #ios .step { text-align:left; font-size:15px; margin:10px 0; color:#333; }
  #ios .sq { display:inline-block; width:22px; height:22px; border:2px solid #2F6FED; border-radius:5px; position:relative; vertical-align:middle; }
  #ios .sq:after { content:"↑"; position:absolute; top:-9px; left:4px; color:#2F6FED; font-weight:700; }
  #ios .close { margin-top:12px; color:#2F6FED; font-weight:700; cursor:pointer; }
</style>
<div id="bar"><button id="installBtn">📲 홈 화면에 설치</button></div>
<div id="ios"><div class="card">
  <div style="font-weight:700;font-size:16px;margin-bottom:6px;">홈 화면에 설치</div>
  <div class="step">① 아래 <b>공유</b> 버튼 <span class="sq"></span> 을 누르세요</div>
  <div class="step">② 목록에서 <b>‘홈 화면에 추가’</b> 를 누르세요</div>
  <div class="close">닫기</div>
</div></div>
<script>
(function(){
  var P = window.parent;
  try {
    var doc = P.document;
    if (!doc.getElementById('pwa-injected')) {
      var m = doc.createElement('meta'); m.id='pwa-injected'; doc.head.appendChild(m);
      function add(tag, attrs){ var el=doc.createElement(tag); for(var k in attrs) el.setAttribute(k, attrs[k]); doc.head.appendChild(el); }
      add('link', {rel:'manifest', href:'/app/static/manifest.json'});
      add('link', {rel:'apple-touch-icon', href:'/app/static/icon-192.png'});
      add('meta', {name:'apple-mobile-web-app-capable', content:'yes'});
      add('meta', {name:'mobile-web-app-capable', content:'yes'});
      add('meta', {name:'apple-mobile-web-app-title', content:'Searching'});
      add('meta', {name:'apple-mobile-web-app-status-bar-style', content:'default'});
      add('meta', {name:'theme-color', content:'#2F6FED'});
    }
    // 옛날 화면 방지: 이전 서비스워커/캐시 제거
    if (P.navigator.serviceWorker) P.navigator.serviceWorker.getRegistrations().then(function(rs){rs.forEach(function(r){r.unregister();});});
    if (P.caches) P.caches.keys().then(function(ks){ks.forEach(function(k){P.caches.delete(k);});});
  } catch(e){}

  var isStandalone = false;
  try { isStandalone = P.matchMedia('(display-mode: standalone)').matches || P.navigator.standalone === true; } catch(e){}
  var ua = ''; try { ua = P.navigator.userAgent || ''; } catch(e){ ua = navigator.userAgent; }
  var isiOS = /iphone|ipad|ipod/i.test(ua);

  var btn = document.getElementById('installBtn');
  var ios = document.getElementById('ios');

  // 안드로이드/크롬: 설치 프롬프트 캡처 (부모 window)
  try {
    if (!P.__deferredPrompt) {
      P.addEventListener('beforeinstallprompt', function(e){ e.preventDefault(); P.__deferredPrompt = e; if(!isStandalone) btn.style.display='inline-block'; });
      P.addEventListener('appinstalled', function(){ btn.style.display='none'; P.__deferredPrompt=null; });
    }
  } catch(e){}

  // 이미 설치됨 → 버튼 숨김. 아니면 표시(설치용 or 안내용)
  if (isStandalone) { btn.style.display='none'; }
  else { btn.style.display='inline-block'; }

  btn.addEventListener('click', async function(){
    if (P.__deferredPrompt) {
      P.__deferredPrompt.prompt();
      try { var r = await P.__deferredPrompt.userChoice; if(r && r.outcome==='accepted') btn.style.display='none'; } catch(e){}
      P.__deferredPrompt = null; return;
    }
    // iOS 또는 프롬프트 미지원 → 안내
    ios.style.display='flex';
  });
  ios.addEventListener('click', function(){ ios.style.display='none'; });

  // 사이드바(햄버거 메뉴) 항목을 고르면 자동으로 닫히게 (모바일)
  try {
    var pdoc = P.document;
    pdoc.addEventListener('click', function(ev){
      var side = pdoc.querySelector('section[data-testid="stSidebar"]');
      if (!side || !side.contains(ev.target)) return;
      if (P.innerWidth > 768) return;               // 모바일에서만
      var lbl = ev.target.closest('label');
      if (!lbl) return;
      setTimeout(function(){
        var c = pdoc.querySelector('[data-testid="stSidebarCollapseButton"] button, [data-testid="stSidebarCollapseButton"], button[aria-label="Close sidebar"]');
        if (c) c.click();
      }, 150);
    }, true);
  } catch(e){}
})();
</script>
        """,
        height=58,
    )


# ──────────────────────────────────────────────────────────────
# 6) 메인 (왼쪽 사이드바 메뉴 = 폰에서 햄버거). 화면은 한 번에 하나만 렌더.
# ──────────────────────────────────────────────────────────────
SCREENS_NAV = [
    ("🔎 검색어 수집", render_search),
    ("📦 주문 수집", render_orders),
    ("🌤️ 날씨", render_weather),
    ("🛒 카페24", render_cafe24),
]


def main():
    # DB 준비 (검색결과·주문·날씨·카페24토큰 테이블 생성)
    try:
        init_db()
    except Exception as e:
        st.error(f"DB 연결 실패: {e}")
        st.stop()

    st.sidebar.markdown("### 🔎 검색기")
    labels = [s[0] for s in SCREENS_NAV]
    choice = st.sidebar.radio("메뉴", labels, label_visibility="collapsed")

    st.title(choice)
    dict(SCREENS_NAV)[choice]()


# ── 앱 시작 ────────────────────────────────────────────────────
mobile_css()
pwa_component()
if not check_password():
    st.stop()
main()
