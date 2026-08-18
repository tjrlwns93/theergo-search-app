# -*- coding: utf-8 -*-
"""
theergo.co.kr 검색 자동화
- 엑셀([검색어] 열)의 각 단어를 더에르고 검색창에 넣고
- 첫 번째 검색 결과의 '제목'과 '링크'를 가져와 새 엑셀로 저장한다.

사용법 (이 폴더에서):
    python theergo_search.py
옵션:
    python theergo_search.py --in "통합 문서1.xlsx" --out "검색결과.xlsx" --col 검색어
"""
import argparse
import html as html_mod
import re
import sys
import time
import urllib.parse
import urllib.request

import openpyxl

BASE = "https://theergo.co.kr"
SEARCH_URL = BASE + "/product/search.html?keyword="
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Search-Bot"}


def fetch(keyword: str) -> str:
    """검색 결과 페이지 HTML을 가져온다."""
    url = SEARCH_URL + urllib.parse.quote(keyword)
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=25) as resp:
        return resp.read().decode("utf-8", "replace")


def strip_tags(fragment: str) -> str:
    """HTML 조각에서 태그를 제거하고 공백을 정리한다."""
    text = re.sub(r"<[^>]+>", " ", fragment)
    text = html_mod.unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    # '상품명 :' 같은 라벨 접두사 제거
    text = re.sub(r"^상품명\s*:\s*", "", text)
    return text


def parse_first_result(page_html: str):
    """검색 결과 페이지에서 첫 번째 상품의 (제목, 링크)를 반환. 없으면 (None, None)."""
    # 상품 목록(prdList) 영역만 대상으로 한다.
    start = page_html.find("prdList")
    if start == -1:
        return None, None
    region = page_html[start:]

    # 첫 번째 상품명 앵커: <strong class="name"><a href="...">제목</a></strong>
    m = re.search(
        r'<strong class="name">\s*<a\s+href="([^"]+)"[^>]*>(.*?)</a>',
        region,
        re.S,
    )
    if not m:
        return None, None

    href = html_mod.unescape(m.group(1).strip())
    title = strip_tags(m.group(2))
    link = urllib.parse.urljoin(BASE, href)
    return (title or None), link


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="infile", default="통합 문서1.xlsx", help="입력 엑셀 파일")
    ap.add_argument("--out", dest="outfile", default="검색결과.xlsx", help="출력 엑셀 파일")
    ap.add_argument("--col", dest="col", default="검색어", help="검색어가 든 열 이름")
    ap.add_argument("--sleep", type=float, default=1.0, help="검색 간 대기(초)")
    args = ap.parse_args()

    # 콘솔 한글 출력 안전 처리
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    # 1) 입력 엑셀 읽기
    wb = openpyxl.load_workbook(args.infile)
    ws = wb.active
    header = [c.value for c in ws[1]]
    if args.col not in header:
        print(f"[오류] '{args.col}' 열을 찾을 수 없습니다. 실제 헤더: {header}")
        sys.exit(1)
    col_idx = header.index(args.col)  # 0-based

    keywords = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[col_idx] if col_idx < len(row) else None
        if val is not None and str(val).strip():
            keywords.append(str(val).strip())

    print(f"검색어 {len(keywords)}개: {keywords}")

    # 2) 출력 엑셀 준비
    out = openpyxl.Workbook()
    ows = out.active
    ows.title = "검색결과"
    ows.append(["검색어", "제목", "링크"])

    # 3) 검색 실행
    for i, kw in enumerate(keywords, 1):
        try:
            page = fetch(kw)
            title, link = parse_first_result(page)
            if title is None:
                title, link = "검색결과 없음", ""
        except Exception as e:
            title, link = f"오류: {e}", ""
        print(f"  [{i}/{len(keywords)}] {kw}  ->  {title} | {link}")
        ows.append([kw, title, link])
        if i < len(keywords):
            time.sleep(args.sleep)

    # 4) 열 너비 보기 좋게
    ows.column_dimensions["A"].width = 14
    ows.column_dimensions["B"].width = 45
    ows.column_dimensions["C"].width = 70

    out.save(args.outfile)
    print(f"\n완료! '{args.outfile}' 에 저장했습니다.")


if __name__ == "__main__":
    main()
